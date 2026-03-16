#!/usr/bin/env python3
# =============================================================================
# ONCOSIEVE — post_pipeline.py
#
# Author : Dr Christopher Trethewey
# Email  : christopher.trethewey@nhs.net
#
# Run after the main pipeline. Performs:
#   1. REVEL annotation on the full whitelist TSV
#   2. High-confidence TSV (removes ClinVar-only zero-sample variants)
#   3. High-confidence VCF (filtered from full VCF via bcftools)
#   4. Excel export of both TSVs with summary sheet and data sheet
#
# Usage:
#   python3 tools/post_pipeline.py \
#       --whitelist output/pan_cancer_whitelist_GRCh38_annotated.tsv.gz \
#       --vcf       output/pan_cancer_whitelist_GRCh38.vcf.gz \
#       --revel     data/REVEL/revel_with_transcript_ids \
#       --out-dir   output/
# =============================================================================

import argparse
import datetime
import importlib.util
import os
import subprocess
import sys
from pathlib import Path

import pandas as pd
import polars as pl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# REVEL annotation
# =============================================================================

def annotate_revel(wl: pd.DataFrame, revel_path: str) -> pd.DataFrame:
    print(f'  Loading REVEL: {revel_path}')
    print('  This may take 1-2 minutes...')
    revel = pl.read_csv(
        revel_path,
        separator=',',
        infer_schema_length=10000,
        null_values=['.'],
        schema_overrides={
            'chr':        pl.Utf8,
            'grch38_pos': pl.Int64,
            'ref':        pl.Utf8,
            'alt':        pl.Utf8,
            'REVEL':      pl.Float64,
        },
    ).select(['chr', 'grch38_pos', 'ref', 'alt', 'REVEL']).filter(
        pl.col('grch38_pos').is_not_null()
    )
    print(f'  {len(revel):,} REVEL entries loaded')

    revel = (
        revel
        .group_by(['chr', 'grch38_pos', 'ref', 'alt'])
        .agg(pl.col('REVEL').max().alias('revel_score'))
    ).to_pandas()
    revel['grch38_pos'] = revel['grch38_pos'].astype('Int64')

    wl['_chrom_key'] = wl['chrom'].str.replace(r'^chr', '', regex=True)
    wl['_pos_key']   = pd.to_numeric(wl['pos'], errors='coerce')

    merged = wl.merge(
        revel,
        left_on  = ['_chrom_key', '_pos_key', 'ref', 'alt'],
        right_on = ['chr', 'grch38_pos', 'ref', 'alt'],
        how      = 'left',
    ).drop(columns=['_chrom_key', '_pos_key', 'chr', 'grch38_pos'])

    merged['revel_score'] = merged['revel_score'].round(4).astype(str)
    merged.loc[merged['revel_score'] == 'nan', 'revel_score'] = ''

    n_scored = (merged['revel_score'] != '').sum()
    print(f'  Variants with REVEL score: {n_scored} / {len(merged)}')
    return merged


# =============================================================================
# Column restructure
# =============================================================================

def restructure_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Extract transcript_id from full hgvsc (e.g. ENST00000288602.6:c.1799T>A)
    # but keep hgvsc as the FULL format (do not strip the transcript prefix)
    import re
    _RE_ENST = re.compile(r'(ENST\d+)', re.I)
    _RE_NM   = re.compile(r'^(N[MRP]_\d+)', re.I)

    def extract_tid(val):
        if not isinstance(val, str) or val in ('', '.'):
            return ''
        m = _RE_ENST.search(val)
        if m:
            return m.group(1).upper()
        m = _RE_NM.match(val)
        if m:
            return m.group(1).upper()
        return ''

    # Fill transcript_id from hgvsc only where not already populated
    if 'transcript_id' not in df.columns:
        df['transcript_id'] = df['hgvsc'].fillna('').apply(extract_tid)
    else:
        missing = df['transcript_id'].fillna('').eq('')
        df.loc[missing, 'transcript_id'] = df.loc[missing, 'hgvsc'].fillna('').apply(extract_tid)

    cols = [
        'chrom', 'pos', 'ref', 'alt', 'gene',
        'transcript_id', 'is_mane_select', 'refseq_id',
        'hgvsc', 'hgvsp', 'protein_change', 'consequence',
        'n_cancer_types', 'cancer_types', 'n_samples', 'sources',
        'oncokb_oncogenicity', 'clinvar_clinical_significance',
        'transcript_source', 'tp53_class', 'wl_tier',
        'genome_version', 'revel_score',
    ]
    extra = [c for c in df.columns if c not in cols]
    df = df[[c for c in cols if c in df.columns] + extra]
    return df


# =============================================================================
# Numeric columns
# =============================================================================

NUMERIC_COLS = ['pos', 'n_cancer_types', 'n_samples', 'wl_tier', 'revel_score']


def coerce_numerics(df: pd.DataFrame) -> pd.DataFrame:
    for col in NUMERIC_COLS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


# =============================================================================
# Summary statistics
# =============================================================================

def compute_summary(df: pd.DataFrame, label: str) -> dict:
    sources       = df['sources'].fillna('').str.split('|').explode().str.strip()
    source_counts = sources[sources != ''].value_counts().to_dict()
    tier_counts   = df['wl_tier'].fillna('').astype(str).value_counts().to_dict()
    csq_counts    = df['consequence'].value_counts().to_dict()
    oncokb_present = df['oncokb_oncogenicity'].fillna('').ne('').sum()
    revel_present  = df['revel_score'].fillna('').astype(str).ne('').sum()
    n_samp = pd.to_numeric(df['n_samples'], errors='coerce').fillna(0)
    clinvar_only   = ((n_samp == 0) &
                      (df['sources'].fillna('') == 'ClinVar')).sum()
    return {
        'label':         label,
        'total':         len(df),
        'tier_counts':   tier_counts,
        'csq_counts':    csq_counts,
        'source_counts': source_counts,
        'oncokb':        int(oncokb_present),
        'revel_scored':  int(revel_present),
        'clinvar_only':  int(clinvar_only),
    }


# =============================================================================
# Styles
# =============================================================================

HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(color='FFFFFF', bold=True, size=10)
SUBHEAD_FILL = PatternFill('solid', fgColor='2E75B6')
SUBHEAD_FONT = Font(color='FFFFFF', bold=True, size=10)
LABEL_FONT   = Font(bold=True, size=10)
THIN         = Side(style='thin', color='CCCCCC')
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TIER_COLORS = {'1': 'C6EFCE', '2': 'FFEB9C', '3': 'FFC7CE'}
CSQ_COLORS  = {
    'missense':      'DEEAF1',
    'nonsense':      'FCE4D6',
    'frameshift':    'FCE4D6',
    'splice_site':   'EBF3FB',
    'splice_region': 'EBF3FB',
    'inframe_indel': 'E2EFDA',
    'synonymous':    'F2F2F2',
}


def _cell(ws, row, col, value='', bold=False, fill=None, font=None,
          align='left', border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font or Font(bold=bold, size=10)
    c.alignment = Alignment(horizontal=align, vertical='center')
    if fill:
        c.fill = fill
    if border:
        c.border = THIN_BORDER
    return c


# =============================================================================
# Summary sheet
# =============================================================================

def write_summary_sheet(wb, stats_full: dict, stats_hc: dict, run_date: str):
    ws = wb.create_sheet('Summary', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value     = 'ONCOSIEVE v1.0 — Pan-Cancer Variant Whitelist'
    c.font      = Font(bold=True, size=14, color='1F4E79')
    c.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A2:C2')
    c = ws['A2']
    c.value     = f'Generated: {run_date}    Author: Dr Christopher Trethewey'
    c.font      = Font(size=9, color='595959')
    c.alignment = Alignment(horizontal='left', vertical='center')

    row = 4

    for col, label in [(1, 'Metric'), (2, 'Full whitelist'),
                       (3, 'High-confidence')]:
        _cell(ws, row, col, label, fill=HEADER_FILL, font=HEADER_FONT,
              align='center')
    row += 1

    def section(title):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=title)
        c.font      = SUBHEAD_FONT
        c.fill      = SUBHEAD_FILL
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border    = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

    def data_row(label, val_full, val_hc):
        nonlocal row
        _cell(ws, row, 1, label, font=LABEL_FONT)
        _cell(ws, row, 2, val_full, align='center')
        _cell(ws, row, 3, val_hc,   align='center')
        ws.row_dimensions[row].height = 15
        row += 1

    section('Overview')
    data_row('Total variants',          stats_full['total'],        stats_hc['total'])
    data_row('ClinVar-only (0 samples)',stats_full['clinvar_only'], stats_hc['clinvar_only'])
    data_row('With OncoKB annotation',  stats_full['oncokb'],       stats_hc['oncokb'])
    data_row('With REVEL score',        stats_full['revel_scored'], stats_hc['revel_scored'])

    section('Tier distribution')
    for tier in ['1', '2', '3']:
        data_row(f'Tier {tier}',
                 stats_full['tier_counts'].get(tier, 0),
                 stats_hc['tier_counts'].get(tier, 0))

    section('Source contribution')
    all_sources = sorted(set(list(stats_full['source_counts'].keys()) +
                             list(stats_hc['source_counts'].keys())))
    for src in all_sources:
        data_row(src,
                 stats_full['source_counts'].get(src, 0),
                 stats_hc['source_counts'].get(src, 0))

    section('Consequence distribution')
    all_csq = sorted(set(list(stats_full['csq_counts'].keys()) +
                          list(stats_hc['csq_counts'].keys())))
    for csq in all_csq:
        data_row(csq,
                 stats_full['csq_counts'].get(csq, 0),
                 stats_hc['csq_counts'].get(csq, 0))

    ws.freeze_panes = 'A5'


# =============================================================================
# Data sheet
# =============================================================================

def write_data_sheet(wb, df: pd.DataFrame, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    cols = list(df.columns)

    col_widths = {
        'chrom': 8, 'pos': 12, 'ref': 6, 'alt': 6, 'gene': 10,
        'transcript_id': 22, 'is_mane_select': 14, 'refseq_id': 18, 'hgvsc': 32,
        'hgvsp': 28, 'protein_change': 18,
        'consequence': 16, 'n_cancer_types': 10, 'cancer_types': 35,
        'n_samples': 10, 'sources': 30, 'oncokb_oncogenicity': 22,
        'clinvar_clinical_significance': 28, 'transcript_source': 16,
        'tp53_class': 16, 'wl_tier': 8,
        'genome_version': 12, 'revel_score': 10,
    }

    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = \
            col_widths.get(col, 14)

    for i, col in enumerate(cols, 1):
        _cell(ws, 1, i, col, fill=HEADER_FILL, font=HEADER_FONT,
              align='center')
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    csq_col = cols.index('consequence') + 1 if 'consequence' in cols else None

    # Pre-build shared style objects to avoid per-cell construction
    data_font = Font(size=9)
    data_align = Alignment(horizontal='left', vertical='center')
    tier_fills = {t: PatternFill('solid', fgColor=c) for t, c in TIER_COLORS.items()}
    csq_fills = {c: PatternFill('solid', fgColor=v) for c, v in CSQ_COLORS.items()}

    # Determine column indices for tier/consequence lookups
    tier_col_idx = cols.index('wl_tier') if 'wl_tier' in cols else None
    csq_col_idx = cols.index('consequence') if 'consequence' in cols else None

    # Batch insert data rows using ws.append (much faster than cell-by-cell)
    for row_vals in df.itertuples(index=False):
        clean = []
        for v in row_vals:
            if isinstance(v, float) and v != v:
                clean.append(None)
            elif v is None or v == 'nan':
                clean.append(None)
            else:
                clean.append(v)
        ws.append(clean)

    # Apply formatting in a second pass (column-oriented where possible)
    n_cols = len(cols)
    for r_idx in range(2, len(df) + 2):
        tier = str(ws.cell(row=r_idx, column=tier_col_idx + 1).value) if tier_col_idx is not None else ''
        csq = str(ws.cell(row=r_idx, column=csq_col_idx + 1).value) if csq_col_idx is not None else ''
        row_fill = tier_fills.get(tier)

        for c_idx in range(1, n_cols + 1):
            c = ws.cell(row=r_idx, column=c_idx)
            c.font = data_font
            c.alignment = data_align
            c.border = THIN_BORDER
            if csq_col is not None and c_idx == csq_col and csq in csq_fills:
                c.fill = csq_fills[csq]
            elif row_fill:
                c.fill = row_fill

        ws.row_dimensions[r_idx].height = 13


# =============================================================================
# Main
# =============================================================================

def main():
    ap = argparse.ArgumentParser(
        description='ONCOSIEVE post-pipeline processing.'
    )
    ap.add_argument('--whitelist', required=True)
    ap.add_argument('--vcf',       required=True)
    ap.add_argument('--revel',     required=True)
    ap.add_argument('--out-dir',   default='output')
    ap.add_argument('--logo',      default=None,
                    help='Path to logo image for HTML report '
                         '(default: assets/logo.jpg relative to repo root)')
    args = ap.parse_args()

    for f in [args.whitelist, args.vcf, args.revel]:
        if not os.path.exists(f):
            sys.exit(f'ERROR: file not found: {f}')

    os.makedirs(args.out_dir, exist_ok=True)
    run_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

    print('\n[1/6] Loading whitelist...')
    wl = pd.read_csv(args.whitelist, sep='\t', dtype=str, low_memory=False)
    print(f'  {len(wl)} variants, {len(wl.columns)} columns')

    print('\n[2/6] Annotating with REVEL scores...')
    wl = annotate_revel(wl, args.revel)

    print('\n[3/6] Restructuring columns...')
    wl = restructure_columns(wl)
    wl = coerce_numerics(wl)
    print(f'  Final columns: {list(wl.columns)}')

    full_tsv = os.path.join(args.out_dir,
                            'pan_cancer_whitelist_GRCh38_full.tsv.gz')
    wl.to_csv(full_tsv, sep='\t', index=False,
              compression='gzip' if full_tsv.endswith('.gz') else None)
    print(f'  Full TSV: {full_tsv}')

    print('\n[4/6] Generating high-confidence whitelist...')
    clinvar_only_mask = (
        (wl['n_samples'].fillna(0) == 0) &
        (wl['sources'].fillna('') == 'ClinVar')
    )
    wl_hc = wl[~clinvar_only_mask].reset_index(drop=True)
    print(f'  Removed {clinvar_only_mask.sum()} ClinVar-only zero-sample variants')
    print(f'  High-confidence variants: {len(wl_hc)}')

    hc_tsv = os.path.join(args.out_dir,
                          'pan_cancer_whitelist_GRCh38_highconf.tsv.gz')
    wl_hc.to_csv(hc_tsv, sep='\t', index=False,
                 compression='gzip' if hc_tsv.endswith('.gz') else None)
    print(f'  High-confidence TSV: {hc_tsv}')

    hc_vcf = os.path.join(args.out_dir,
                          'pan_cancer_whitelist_GRCh38_highconf.vcf.gz')
    print('  Filtering VCF with bcftools...')
    cmd = (
        f"bcftools filter -e "
        f"\"INFO/N_SAMPLES=0 && INFO/SOURCES='ClinVar'\" "
        f"{args.vcf} -O z -o {hc_vcf} && tabix -f -p vcf {hc_vcf}"
    )
    result = subprocess.run(cmd, shell=True)
    if result.returncode != 0:
        print('  WARNING: bcftools filtering failed.')
    else:
        print(f'  High-confidence VCF: {hc_vcf}')

    print('\n[5/6] Exporting Excel files...')
    stats_full = compute_summary(wl,    'Full whitelist')
    stats_hc   = compute_summary(wl_hc, 'High-confidence')

    full_xlsx = os.path.join(args.out_dir,
                             'pan_cancer_whitelist_GRCh38_full.xlsx')
    wb_full = openpyxl.Workbook()
    wb_full.remove(wb_full.active)
    write_summary_sheet(wb_full, stats_full, stats_hc, run_date)
    write_data_sheet(wb_full, wl,    'Full whitelist')
    wb_full.save(full_xlsx)
    print(f'  Full xlsx: {full_xlsx}')

    hc_xlsx = os.path.join(args.out_dir,
                           'pan_cancer_whitelist_GRCh38_highconf.xlsx')
    wb_hc = openpyxl.Workbook()
    wb_hc.remove(wb_hc.active)
    write_summary_sheet(wb_hc, stats_full, stats_hc, run_date)
    write_data_sheet(wb_hc, wl_hc, 'High-confidence')
    wb_hc.save(hc_xlsx)
    print(f'  High-confidence xlsx: {hc_xlsx}')

    print('\n[6/6] Generating HTML report...')
    _repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _report_script = os.path.join(_repo_root, 'tools', 'generate_report.py')
    if not os.path.exists(_report_script):
        print(f'  WARNING: generate_report.py not found at {_report_script}, skipping.')
        report_path = None
    else:
        # Resolve logo: use --logo arg, else default to assets/logo.jpg in repo root
        logo_path = args.logo
        if logo_path is None:
            _default_logo = os.path.join(_repo_root, 'assets', 'logo.jpg')
            logo_path = _default_logo if os.path.exists(_default_logo) else None
        if logo_path:
            print(f'  Logo: {logo_path}')
        else:
            print('  No logo found at assets/logo.jpg, report will have no logo.')

        # Load generate_report module dynamically (avoids sys.path manipulation)
        spec = importlib.util.spec_from_file_location('generate_report', _report_script)
        gr   = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(gr)

        report_date = datetime.datetime.now().strftime('%Y-%m-%d')
        report_path = os.path.join(args.out_dir,
                                   f'oncosieve_report_{report_date}.html')
        gr.build_report(
            wl,
            wl_hc,
            Path(report_path),
            logo_path=logo_path,
        )

    print('\nPost-pipeline complete.')
    print(f'  Full TSV:             {full_tsv}')
    print(f'  High-confidence TSV:  {hc_tsv}')
    print(f'  High-confidence VCF:  {hc_vcf}')
    print(f'  Full xlsx:            {full_xlsx}')
    print(f'  High-confidence xlsx: {hc_xlsx}')
    if report_path and os.path.exists(report_path):
        print(f'  HTML report:          {report_path}')


if __name__ == '__main__':
    main()
